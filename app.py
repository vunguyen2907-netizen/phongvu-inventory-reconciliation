import streamlit as st
import pandas as pd
import io
import datetime
import os
import re
import uuid
from supabase_store import (
    dataframe_from_payload,
    display_time,
    download_source_file,
    is_configured as supabase_is_configured,
    list_inventory_sessions,
    load_inventory_session,
    save_inventory_session,
    upload_source_files,
)

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as ReportlabImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

st.set_page_config(page_title="Đối Soát Kiểm Kê - Phong Vũ", layout="wide")

# Inject Custom Minimalist Corporate CSS
st.markdown("""
<style>
    html, body, [class*="css"] {
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    }
    .app-header {
        padding: 0.2rem 0 1.2rem 0;
        border-bottom: 1px solid #E5E7EB;
        margin-bottom: 1.2rem;
    }
    .app-title {
        font-size: 1.5rem;
        font-weight: 700;
        color: #003399;
        margin: 0;
        padding: 0;
        letter-spacing: -0.01em;
    }
    .app-subtitle {
        font-size: 0.88rem;
        color: #6B7280;
        margin-top: 0.25rem;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 6px;
        border-bottom: 1px solid #E5E7EB;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 16px;
        font-weight: 500;
        font-size: 0.88rem;
        color: #4B5563;
        border-radius: 4px 4px 0 0;
    }
    .stTabs [aria-selected="true"] {
        color: #003399 !important;
        font-weight: 600 !important;
        border-bottom: 2px solid #003399 !important;
    }
    [data-testid="stMetricValue"] {
        font-size: 1.35rem !important;
        font-weight: 700 !important;
        color: #111827 !important;
    }
    [data-testid="stMetricLabel"] {
        font-size: 0.82rem !important;
        font-weight: 500 !important;
        color: #6B7280 !important;
    }
    .stButton>button {
        border-radius: 4px;
        font-weight: 500;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="app-header">
    <div class="app-title">Hệ thống Tự động Xử lý & Đối soát Dữ liệu Kiểm kê</div>
    <div class="app-subtitle">Công ty Cổ phần Thương mại Dịch vụ Phong Vũ — Quy trình đối soát & xuất báo cáo định kỳ</div>
</div>
""", unsafe_allow_html=True)

# Đăng ký Font Tiếng Việt cho ReportLab PDF
pdf_font_path = '/System/Library/Fonts/Supplemental/Arial.ttf'
if not os.path.exists(pdf_font_path):
    pdf_font_path = '/Library/Fonts/Arial.ttf'

if os.path.exists(pdf_font_path):
    try:
        pdfmetrics.registerFont(TTFont('ArialVN', pdf_font_path))
    except Exception:
        pass

def find_column(columns, candidates):
    """
    Tìm tên cột phù hợp nhất theo danh sách từ khóa ưu tiên.
    """
    cols_clean = {c: str(c).strip().lower() for c in columns if pd.notna(c)}
    for kw in candidates:
        for orig_col, clean_col in cols_clean.items():
            if kw in clean_col:
                return orig_col
    return None

def normalize_serial(s):
    """
    Chuẩn hóa Serial đối soát chính xác:
    - Viết hoa toàn bộ
    - Loại bỏ khoảng trắng thừa và ký tự ẩn unicode
    """
    if pd.isna(s):
        return ""
    st_val = str(s).strip().upper()
    return "".join(c for c in st_val if c.isalnum() or c in ['-', '_', '/', '.'])

def clean_export_note(note_str):
    """
    Loại bỏ các thẻ ghi chú nội bộ đếm lần 2 như '[Bổ sung Serial Lần 2: ...]', '[Sửa Serial: ...]', '[Đã loại bỏ Serial...]'
    Chỉ giữ lại các mã đơn hàng xuất bán và tình trạng dư/thiếu thực tế cho báo cáo gửi sếp.
    """
    if pd.isna(note_str):
        return ""
    s = str(note_str).strip()
    if not s:
        return ""
    
    s_clean = re.sub(r'\[Bổ sung Serial[^\]]*\]', '', s)
    s_clean = re.sub(r'\[Sửa Serial[^\]]*\]', '', s_clean)
    s_clean = re.sub(r'\[Đã loại bỏ Serial[^\]]*\]', '', s_clean)
    
    parts = [p.strip() for p in s_clean.split('|') if p.strip()]
    return " | ".join(parts)

def is_format_mismatch(scanned_str, stock_list):
    """
    Kiểm tra xem serial được quét (scanned_str) có bị lệch định dạng so với danh sách serial sổ sách (stock_list) không:
    - Lệch độ dài (length mismatch >= 2)
    - Lệch cấu trúc chữ/số (stock có chứa chữ nhưng scanned lại là thuần số, hoặc ngược lại)
    - Lệch tiền tố chung (nếu các serial trong sổ sách có 2 ký tự đầu giống hệt nhau)
    """
    if not stock_list:
        return False
    
    scanned_str = str(scanned_str).strip()
    if not scanned_str or scanned_str in ['nan', 'None', 'null', '-']:
        return False
    
    clean_stock = [str(s).strip() for s in stock_list if pd.notna(s) and str(s).strip() not in ['', 'nan', 'None', 'null', '-']]
    if not clean_stock:
        return False

    norm_scanned = normalize_serial(scanned_str)
    norm_stock = [normalize_serial(s) for s in clean_stock if normalize_serial(s)]

    stock_lengths = set(len(s) for s in norm_stock)
    has_letters_in_stock = any(any(c.isalpha() for c in s) for s in clean_stock)
    prefixes = set(s[:2].upper() for s in clean_stock if len(s) >= 2)

    scanned_len = len(norm_scanned)
    scanned_has_letters = any(c.isalpha() for c in scanned_str)
    scanned_prefix = scanned_str[:2].upper() if len(scanned_str) >= 2 else ""

    # 1. Lệch độ dài đáng kể
    if scanned_len not in stock_lengths and min(abs(scanned_len - l) for l in stock_lengths) >= 2:
        return True

    # 2. Lệch loại ký tự
    if has_letters_in_stock and not scanned_has_letters:
        return True
    if not has_letters_in_stock and scanned_has_letters:
        return True

    # 3. Lệch tiền tố chung
    if len(prefixes) == 1 and scanned_prefix and scanned_prefix not in prefixes:
        return True

    return False


def read_smart_dataframe(file_obj, is_erp=False):
    """
    Đọc dữ liệu từ file CSV hoặc Excel:
    - Nếu là file ERP Excel: Tự động chọn Sheet 2 ('KetQuaChiTiet' hoặc sheet index 1).
    - Tự động dò tìm dòng Tiêu đề (Header) bằng cách chấm điểm các từ khóa tiêu đề phổ biến.
    """
    filename = file_obj.name.lower()
    
    if filename.endswith('.csv'):
        file_obj.seek(0)
        df_raw = pd.read_csv(file_obj, header=None)
    else:
        file_obj.seek(0)
        target_sheet = 0
        if is_erp:
            try:
                try:
                    xl = pd.ExcelFile(file_obj, engine='calamine')
                except Exception:
                    xl = pd.ExcelFile(file_obj, engine='openpyxl')
                
                sheets = xl.sheet_names
                if 'KetQuaChiTiet' in sheets:
                    target_sheet = 'KetQuaChiTiet'
                elif len(sheets) > 1:
                    target_sheet = 1
            except Exception:
                target_sheet = 0
        
        file_obj.seek(0)
        try:
            df_raw = pd.read_excel(file_obj, sheet_name=target_sheet, header=None, engine='calamine')
        except Exception:
            df_raw = pd.read_excel(file_obj, sheet_name=target_sheet, header=None, engine='openpyxl')

    header_keywords = ['mã', 'sku', 'tên', 'tồn', 'số lượng', 'sl', 'serial', 'chênh lệch', 'đvt', 'hàng hóa', 'sản phẩm', 'stt', 'đơn vị', 'part']
    
    best_row_idx = 0
    max_score = -1
    
    for idx in range(min(30, len(df_raw))):
        row_cells = [str(val).strip() for val in df_raw.iloc[idx].values if pd.notna(val) and str(val).strip() != '']
        row_cells_lower = [c.lower() for c in row_cells]
        
        distinct_cells = set(row_cells_lower)
        if len(distinct_cells) < 3:
            continue
        
        score = sum(1 for cell in distinct_cells if any(kw in cell for kw in header_keywords))
        if score > max_score:
            max_score = score
            best_row_idx = idx

    header_raw = df_raw.iloc[best_row_idx].values
    seen = {}
    clean_headers = []
    for i, val in enumerate(header_raw):
        base = str(val).strip() if pd.notna(val) and str(val).strip() != '' else f"Unnamed_{i}"
        if base in seen:
            seen[base] += 1
            clean_headers.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 0
            clean_headers.append(base)

    df = df_raw.iloc[best_row_idx + 1:].copy().reset_index(drop=True)
    df.columns = clean_headers
    return df


def save_session_state(show_progress=False):
    """Tự động lưu đợt kiểm kê hiện tại lên Supabase nếu đã cấu hình."""
    if not supabase_is_configured() or st.session_state.df_check_detail is None:
        return False

    def persist():
        try:
            saved = save_inventory_session(
                st.session_state.get('active_session_id'),
                st.session_state.get('session_name', ''),
                st.session_state.get('df_recon'),
                st.session_state.get('df_count_l2'),
                st.session_state.get('df_check_detail'),
                st.session_state.get('source_files'),
            )
            st.session_state.active_session_id = saved['id']

            # File ERP và tồn kho gốc không nằm trong JSONB; chúng được đưa
            # vào Supabase Storage private đúng một lần cho từng đợt.
            pending_files = st.session_state.get('pending_source_uploads', [])
            if pending_files:
                upload_source_files(pending_files)
                st.session_state.pending_source_uploads = []

            list_inventory_sessions.clear()
            st.session_state.last_saved_time = display_time(saved.get('updated_at'))
            st.session_state.db_error = None
            st.session_state.save_notice = "Đã lưu dữ liệu kiểm kê lên cloud."
            return True
        except Exception as exc:
            st.session_state.db_error = str(exc)
            return False

    if show_progress:
        with st.spinner("Đang lưu và đồng bộ dữ liệu lên cloud…"):
            return persist()
    return persist()

def reset_session_state():
    """Dọn màn hình để tạo một đợt mới; lịch sử Supabase vẫn được giữ."""
    st.session_state.df_recon = None
    st.session_state.df_count_l2 = None
    st.session_state.df_check_detail = None
    st.session_state.pending_clear_serial = None
    st.session_state.source_files = []
    st.session_state.pending_source_uploads = []
    st.session_state.source_file_download = None
    st.session_state.active_session_id = None
    st.session_state.session_name = f"Kiểm kê {datetime.date.today().strftime('%d/%m/%Y')}"
    st.session_state.last_saved_time = None

# Khởi tạo session state & tự động khôi phục dữ liệu khi F5
if 'df_count_l2' not in st.session_state:
    st.session_state.df_count_l2 = None
if 'df_check_detail' not in st.session_state:
    st.session_state.df_check_detail = None
if 'df_recon' not in st.session_state:
    st.session_state.df_recon = None
if 'pending_clear_serial' not in st.session_state:
    st.session_state.pending_clear_serial = None
if 'active_session_id' not in st.session_state:
    st.session_state.active_session_id = None
if 'session_name' not in st.session_state:
    st.session_state.session_name = f"Kiểm kê {datetime.date.today().strftime('%d/%m/%Y')}"
if 'last_saved_time' not in st.session_state:
    st.session_state.last_saved_time = None
if 'db_error' not in st.session_state:
    st.session_state.db_error = None
if 'source_files' not in st.session_state:
    st.session_state.source_files = []
if 'pending_source_uploads' not in st.session_state:
    st.session_state.pending_source_uploads = []
if 'save_notice' not in st.session_state:
    st.session_state.save_notice = None
if 'source_file_download' not in st.session_state:
    st.session_state.source_file_download = None

def load_selected_session(session_id):
    saved = load_inventory_session(session_id)
    payload = saved.get('data', {})
    st.session_state.df_recon = dataframe_from_payload(payload.get('df_recon'))
    st.session_state.df_count_l2 = dataframe_from_payload(payload.get('df_count_l2'))
    st.session_state.df_check_detail = dataframe_from_payload(payload.get('df_check_detail'))
    st.session_state.source_files = payload.get('source_files', [])
    st.session_state.pending_source_uploads = []
    st.session_state.source_file_download = None
    st.session_state.active_session_id = saved['id']
    st.session_state.session_name = saved.get('session_name', '')
    st.session_state.last_saved_time = display_time(saved.get('updated_at'))


with st.sidebar:
    st.header("Lịch sử kiểm kê")
    st.text_input("Tên đợt kiểm kê", key="session_name")
    if supabase_is_configured():
        if st.button("Lưu đợt hiện tại", use_container_width=True, disabled=st.session_state.df_check_detail is None):
            save_session_state(show_progress=True)
            st.rerun()
        try:
            saved_sessions = list_inventory_sessions()
            session_ids = [row['id'] for row in saved_sessions]
            if session_ids:
                selected_id = st.selectbox(
                    "Mở lịch sử",
                    session_ids,
                    format_func=lambda item: next(
                        f"{row['session_name']} — {display_time(row.get('updated_at'))}"
                        for row in saved_sessions if row['id'] == item
                    ),
                )
                if st.button("Mở đợt đã chọn", use_container_width=True):
                    load_selected_session(selected_id)
                    st.rerun()
            else:
                st.caption("Chưa có đợt kiểm kê nào được lưu.")
        except Exception as exc:
            st.error(f"Không tải được lịch sử Supabase: {exc}")
    else:
        st.warning("Chưa kết nối Supabase. Xem README để cấu hình secrets.")

    if st.button("Tạo đợt kiểm kê mới", use_container_width=True):
        reset_session_state()
        st.rerun()

    if st.session_state.last_saved_time:
        st.caption(f"Đã lưu: {st.session_state.last_saved_time}")
    if st.session_state.source_files:
        st.caption("Đã lưu file nguồn: " + ", ".join(item['name'] for item in st.session_state.source_files))
        with st.expander("Tải lại dữ liệu nguồn của đợt này"):
            for source in st.session_state.source_files:
                source_id = source['path'].replace('/', '_').replace('.', '_')
                if st.button(f"Chuẩn bị tải: {source['name']}", key=f"prepare_source_{source_id}", use_container_width=True):
                    try:
                        with st.spinner(f"Đang tải {source['name']}…"):
                            st.session_state.source_file_download = {
                                'path': source['path'],
                                'name': source['name'],
                                'content_type': source.get('content_type', 'application/octet-stream'),
                                'data': download_source_file(source['path']),
                            }
                    except Exception as exc:
                        st.error(f"Không tải được file nguồn: {exc}")
                cached_download = st.session_state.source_file_download
                if cached_download and cached_download['path'] == source['path']:
                    st.download_button(
                        label=f"Tải xuống {source['name']}",
                        data=cached_download['data'],
                        file_name=source['name'],
                        mime=cached_download['content_type'],
                        key=f"download_source_{source_id}",
                        use_container_width=True,
                    )
    if st.session_state.save_notice:
        st.success(st.session_state.save_notice)
        st.session_state.save_notice = None
    if st.session_state.db_error:
        st.warning(f"Chưa lưu được lên Supabase: {st.session_state.db_error}")

@st.dialog("Lý Do Loại Bỏ Serial Bắn Dư Khỏi Báo Cáo")
def modal_clear_surplus(sku_val, serial_val):
    st.markdown(f"Bạn đang yêu cầu **loại bỏ Serial bắn dư `{serial_val}`** của SKU **`{sku_val}`** khỏi Báo cáo Tổng hợp hàng hóa (`TH-HANG HOA`).")
    reason_input = st.text_input("Vui lòng nhập lý do loại bỏ (Ví dụ: Quét nhầm, hàng mẫu, hủy đơn xuất...):", key="reason_clear_input")
    
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        if st.button("Xác nhận loại bỏ khỏi Báo cáo", type="primary", use_container_width=True):
            if not reason_input.strip():
                st.warning("Bạn phải nhập lý do loại bỏ trước khi tiếp tục.")
            else:
                mask = (st.session_state.df_check_detail['Mã sản phẩm'] == sku_val) & (
                    (st.session_state.df_check_detail['Serial đã quét'] == serial_val) |
                    (st.session_state.df_check_detail['Note đơn hàng'].str.contains(serial_val, na=False))
                )
                st.session_state.df_check_detail.loc[mask, 'Serial đã quét'] = ""
                st.session_state.df_check_detail.loc[mask, 'Đã kiểm'] = 0
                st.session_state.df_check_detail.loc[mask, 'Dư/Thiếu'] = 0
                st.session_state.df_check_detail.loc[mask, 'Check đơn'] = "Đã loại bỏ Serial dư"
                
                history_tag = f"[Đã loại bỏ Serial dư {serial_val} - Lý do: {reason_input.strip()}]"
                st.session_state.df_check_detail.loc[mask, 'Note đơn hàng'] = history_tag

                l2_mask = st.session_state.df_count_l2['Mã vật tư'] == sku_val
                if (st.session_state.df_count_l2.loc[l2_mask, 'Số liệu thực tế đã đếm lần 1'] > 0).any():
                    st.session_state.df_count_l2.loc[l2_mask, 'Số liệu thực tế đã đếm lần 1'] -= 1

                st.session_state.pending_clear_serial = None
                save_session_state(show_progress=True)
                st.success(f"Đã loại bỏ Serial {serial_val}!")
                st.rerun()
    with col_c2:
        if st.button("Hủy thao tác", use_container_width=True):
            st.session_state.pending_clear_serial = None
            st.rerun()

tabs = st.tabs([
    "1. Nhập dữ liệu",
    "2. Kiểm đếm lần 2",
    "3. Kết quả kiểm lần 2",
    "4. Bảng tổng hợp chênh lệch",
    "5. Xuất báo cáo"
])

# ==========================================
# TAB 1: NHẬP DỮ LIỆU ĐẦU VÀO
# ==========================================
with tabs[0]:
    st.subheader("Tải lên dữ liệu hệ thống gốc & kết quả kiểm đếm thực tế")
    
    if st.session_state.last_saved_time and st.session_state.df_check_detail is not None:
        st.info(f"Đợt kiểm kê đang mở đã được lưu lúc: {st.session_state.last_saved_time}")

    col1, col2 = st.columns(2)
    
    with col1:
        with st.container(border=True):
            st.markdown("**Dữ liệu Tồn kho Sổ sách (Hệ thống gốc)**")
            file_stock = st.file_uploader("Chọn file Tồn kho gốc (Chứa SKU, Số lượng, Serial, Bin...)", type=["csv", "xlsx"], key="stock")
        
    with col2:
        with st.container(border=True):
            st.markdown("**Dữ liệu Kiểm đếm Thực tế (ERP)**")
            file_erp = st.file_uploader("Chọn file Kết quả kiểm đếm thực tế từ ERP...", type=["csv", "xlsx"], key="erp")

    if file_stock and file_erp:
        st.info("Đã nhận đủ dữ liệu hệ thống và ERP. Nhấn nút bên dưới để bắt đầu tự động trích xuất & đối soát.")
        if st.button("Bắt đầu trích xuất & Tạo 3 bảng đối soát", type="primary", use_container_width=True):
            # Mỗi lần import là một đợt mới. File gốc sẽ được lưu ở Storage,
            # còn các bảng đối soát lưu ở PostgreSQL/JSONB.
            new_session_id = str(uuid.uuid4())
            stock_bytes = file_stock.getvalue()
            erp_bytes = file_erp.getvalue()
            def source_record(file_kind, uploaded_file, content):
                extension = os.path.splitext(uploaded_file.name)[1].lower()
                return {
                    'kind': file_kind,
                    'name': uploaded_file.name,
                    'path': f"{new_session_id}/{file_kind}{extension}",
                    'content_type': uploaded_file.type or 'application/octet-stream',
                    'size': len(content),
                }
            stock_source = source_record('ton-kho-goc', file_stock, stock_bytes)
            erp_source = source_record('kiem-dem-erp', file_erp, erp_bytes)
            st.session_state.active_session_id = new_session_id
            st.session_state.source_files = [stock_source, erp_source]
            st.session_state.pending_source_uploads = [
                {**stock_source, 'content': stock_bytes},
                {**erp_source, 'content': erp_bytes},
            ]
            
            df_stock_raw = read_smart_dataframe(file_stock, is_erp=False)
            df_erp_raw = read_smart_dataframe(file_erp, is_erp=True)
            
            sku_candidates = ['mã sản phẩm', 'mã hàng hóa', 'mã hàng', 'mã sp', 'mã hh', 'mã vật tư', 'mã vt', 'mã thiết bị', 'sku', 'item code', 'product code', 'mã']
            name_candidates = ['tên sản phẩm', 'tên hàng hóa', 'tên hàng', 'tên sp', 'tên hh', 'tên vật tư', 'tên thiết bị', 'description', 'product name', 'item name', 'tên']
            uom_candidates = ['đơn vị tính', 'đvt', 'uom', 'unit', 'đơn vị']
            qty_candidates = ['số lượng sổ sách', 'tồn cuối', 'số lượng tồn', 'tồn kho', 'tồn sổ sách', 'sl sổ sách', 'sl tồn', 'tồn cuối kỳ', 'số lượng', 'tồn', 'sl']
            serial_candidates = ['serial/lot', 'serial', 'imei', 'sn', 'số serial', 'seri']
            part_candidates = ['part number', 'part_number', 'mã part', 'part no', 'part', 'mã phụ tùng']
            bin_candidates = ['mã bin', 'tên bin', 'bin kiểm đếm', 'bin', 'vị trí', 'khu vực', 'kệ', 'ô kiểm đếm']
            type_candidates = ['loại hàng', 'loại serial', 'loại', 'hàng hóa', 'phân loại']

            erp_sku_candidates = sku_candidates
            erp_serial_candidates = serial_candidates
            erp_qty_candidates = ['số lượng thực tế', 'thực tế đếm', 'thực tế đã đếm', 'sl thực tế', 'đã đếm', 'số lượng kiểm đếm', 'sl kiểm đếm', 'số lượng', 'sl']

            sku_col = find_column(df_stock_raw.columns, sku_candidates)
            name_col = find_column(df_stock_raw.columns, name_candidates)
            uom_col = find_column(df_stock_raw.columns, uom_candidates)
            qty_col = find_column(df_stock_raw.columns, qty_candidates)
            serial_col = find_column(df_stock_raw.columns, serial_candidates)
            part_col = find_column(df_stock_raw.columns, part_candidates)
            bin_col = find_column(df_stock_raw.columns, bin_candidates)
            type_col = find_column(df_stock_raw.columns, type_candidates)
            
            erp_sku_col = find_column(df_erp_raw.columns, erp_sku_candidates)
            erp_serial_col = find_column(df_erp_raw.columns, erp_serial_candidates)
            erp_qty_col = find_column(df_erp_raw.columns, erp_qty_candidates)

            if not sku_col or not qty_col:
                st.error("Không tìm thấy tự động cột Mã sản phẩm hoặc Số lượng tồn trên file Hệ thống gốc.")
            elif not erp_sku_col:
                st.error("Không tìm thấy cột Mã sản phẩm (SKU) trên file ERP.")
            else:
                # Bỏ qua các dòng tổng cộng summary cuối file
                invalid_skus = ['tổng', 'tổng cộng', 'total', 'nan', 'none', 'null', 'sum', '']
                
                # 1. Chuẩn hóa dữ liệu sổ sách
                df_stock_clean = pd.DataFrame()
                df_stock_clean['SKU'] = df_stock_raw[sku_col].astype(str).str.strip()
                df_stock_clean['Tên sản phẩm'] = df_stock_raw[name_col] if name_col else "Chưa có tên"
                df_stock_clean['ĐVT'] = df_stock_raw[uom_col] if uom_col else "Cái"
                df_stock_clean['Số lượng sổ sách'] = pd.to_numeric(df_stock_raw[qty_col], errors='coerce').fillna(0).astype(int)
                
                # Lọc bỏ dòng tổng cộng
                mask_stock = (
                    df_stock_clean['SKU'].notna() &
                    (~df_stock_clean['SKU'].str.lower().isin(invalid_skus)) &
                    (~df_stock_clean['ĐVT'].astype(str).str.strip().str.lower().isin(['tổng', 'tổng cộng', 'total'])) &
                    (~df_stock_clean['Tên sản phẩm'].astype(str).str.strip().str.lower().isin(['tổng cộng', 'tổng số lượng', 'total']))
                )
                df_stock_clean = df_stock_clean[mask_stock].copy()
                df_stock_valid_raw = df_stock_raw.loc[df_stock_clean.index].copy()

                # Gom nhóm theo SKU cho sổ sách
                df_stock_agg = df_stock_clean.groupby('SKU').agg({
                    'Tên sản phẩm': 'first',
                    'ĐVT': 'first',
                    'Số lượng sổ sách': 'sum'
                }).reset_index()

                # -------------------------------------------------------------
                # 0. SÀNG LỌC & GẮN CỜ ĐỊNH DANH (SERIAL VS NON-SERIAL) TỪ TỒN KHO GỐC
                # -------------------------------------------------------------
                serial_skus = set()
                if serial_col:
                    for _, row in df_stock_valid_raw.iterrows():
                        sku_val = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
                        s_val = str(row[serial_col]).strip() if pd.notna(row[serial_col]) else ""
                        if sku_val and s_val and s_val.lower() not in ['', 'nan', 'none', 'null', '-']:
                            serial_skus.add(sku_val)

                # 2. Chuẩn hóa dữ liệu ERP thực tế
                df_erp_raw[erp_sku_col] = df_erp_raw[erp_sku_col].astype(str).str.strip()
                mask_erp = (
                    df_erp_raw[erp_sku_col].notna() &
                    (~df_erp_raw[erp_sku_col].str.lower().isin(invalid_skus))
                )
                df_erp_clean = df_erp_raw[mask_erp].copy()

                # Tách riêng số liệu đếm ERP cho SKU Non-Serial và SKU Serial
                erp_qty_dict = {}
                if erp_qty_col:
                    df_erp_clean[erp_qty_col] = pd.to_numeric(df_erp_clean[erp_qty_col], errors='coerce').fillna(0)
                    df_erp_valid = df_erp_clean[df_erp_clean[erp_qty_col] > 0].copy()
                    erp_qty_dict = df_erp_valid.groupby(erp_sku_col)[erp_qty_col].sum().to_dict()

                # Lấy danh sách Serial ERP cho các SKU có quản lý Serial
                erp_serials_by_sku = {}
                if erp_serial_col:
                    if erp_qty_col:
                        df_erp_serial_rows = df_erp_clean[(df_erp_clean[erp_qty_col] > 0) & (df_erp_clean[erp_serial_col].notna())].copy()
                    else:
                        df_erp_serial_rows = df_erp_clean[df_erp_clean[erp_serial_col].notna()].copy()
                        
                    for _, erp_r in df_erp_serial_rows.iterrows():
                        s_sku = str(erp_r[erp_sku_col]).strip() if pd.notna(erp_r[erp_sku_col]) else ""
                        s_ser = str(erp_r[erp_serial_col]).strip() if pd.notna(erp_r[erp_serial_col]) else ""
                        if s_sku and s_ser and s_ser.lower() not in ['', 'nan', 'none', 'null', '-']:
                            if s_sku not in erp_serials_by_sku:
                                erp_serials_by_sku[s_sku] = []
                            erp_serials_by_sku[s_sku].append(s_ser)

                # Tổng hợp số lượng thực tế ERP cho Bảng tổng quan (df_recon)
                erp_recon_dict = {}
                all_skus = set(df_stock_agg['SKU']).union(set(df_erp_clean[erp_sku_col]))
                for s_item in all_skus:
                    if s_item in serial_skus:
                        erp_recon_dict[s_item] = len(erp_serials_by_sku.get(s_item, []))
                    else:
                        erp_recon_dict[s_item] = int(erp_qty_dict.get(s_item, 0))

                df_erp_agg = pd.DataFrame(list(erp_recon_dict.items()), columns=['SKU', 'Số lượng thực tế ERP'])

                df_recon = pd.merge(df_stock_agg, df_erp_agg, on='SKU', how='outer')
                df_recon['Số lượng thực tế ERP'] = df_recon['Số lượng thực tế ERP'].fillna(0).astype(int)
                df_recon['Số lượng sổ sách'] = df_recon['Số lượng sổ sách'].fillna(0).astype(int)
                df_recon['Tên sản phẩm'] = df_recon['Tên sản phẩm'].fillna("Sản phẩm mới chưa có trong danh mục")
                df_recon['ĐVT'] = df_recon['ĐVT'].fillna("Cái")
                st.session_state.df_recon = df_recon

                # -------------------------------------------------------------
                # TẠO BẢNG 1: BẢNG KIỂM ĐẾM LẦN 2
                # -------------------------------------------------------------
                df_l2 = pd.DataFrame()
                df_l2['Mã vật tư'] = df_recon['SKU']
                df_l2['Tên vật tư'] = df_recon['Tên sản phẩm']
                df_l2['Đvt'] = df_recon['ĐVT']
                df_l2['Số lượng sổ sách (hệ thống)'] = df_recon['Số lượng sổ sách']
                df_l2['Số liệu thực tế đã đếm lần 1'] = df_recon['Số lượng thực tế ERP']
                df_l2['Số liệu thực tế đếm lại lần 2'] = 0
                df_l2['Chênh lệch sau kiểm đếm lần 2'] = (df_l2['Số liệu thực tế đã đếm lần 1'] + df_l2['Số liệu thực tế đếm lại lần 2']) - df_l2['Số lượng sổ sách (hệ thống)']
                df_l2['Kết quả xử lý sau khi đếm lại lần 2'] = df_l2['Số liệu thực tế đã đếm lần 1'] + df_l2['Số liệu thực tế đếm lại lần 2']
                df_l2['Note mã đơn'] = ""
                st.session_state.df_count_l2 = df_l2

                # -------------------------------------------------------------
                # TẠO BẢNG 3: BẢNG CHI TIẾT SERIAL THEO BIN (Sheet 'check')
                # -------------------------------------------------------------
                detail_rows = []
                stock_rows_by_sku = {}
                
                for idx_row, row in df_stock_valid_raw.iterrows():
                    sku_val = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
                    if not sku_val or sku_val.lower() in invalid_skus:
                        continue

                    name_val = str(row[name_col]).strip() if name_col and pd.notna(row[name_col]) else "Chưa có tên"
                    part_val = str(row[part_col]).strip() if part_col and pd.notna(row[part_col]) else ""
                    uom_val = str(row[uom_col]).strip() if uom_col and pd.notna(row[uom_col]) else "Cái"
                    serial_val = str(row[serial_col]).strip() if serial_col and pd.notna(row[serial_col]) else ""
                    bin_val = str(row[bin_col]).strip() if bin_col and pd.notna(row[bin_col]) else ""
                    type_val = str(row[type_col]).strip() if type_col and pd.notna(row[type_col]) else "Hàng bán"

                    row_qty = int(pd.to_numeric(row[qty_col], errors='coerce')) if qty_col and pd.notna(row[qty_col]) else 1
                    if row_qty <= 0:
                        row_qty = 1

                    if sku_val not in stock_rows_by_sku:
                        stock_rows_by_sku[sku_val] = []

                    is_non_serial = sku_val not in serial_skus
                    actual_qty = row_qty if is_non_serial else 1

                    stock_rows_by_sku[sku_val].append({
                        "Mã sản phẩm": sku_val,
                        "Tên sản phẩm": name_val,
                        "Part number": part_val,
                        "ĐVT": uom_val,
                        "Số lượng": actual_qty,
                        "Số Serial": "(Hàng không serial)" if is_non_serial else serial_val,
                        "Serial đã quét": "",
                        "Mã Bin": bin_val,
                        "Tên Bin": bin_val,
                        "Loại hàng": "Hàng không serial" if is_non_serial else type_val,
                        "Đã kiểm": 0,
                        "Dư/Thiếu": -actual_qty,
                        "Check đơn": "Bắn thiếu (Chưa quét)",
                        "Note đơn hàng": "",
                        "is_non_serial": is_non_serial
                    })

                all_skus_set = set(stock_rows_by_sku.keys()).union(set(erp_serials_by_sku.keys())).union(set(erp_qty_dict.keys()))

                for sku_val in all_skus_set:
                    s_rows = stock_rows_by_sku.get(sku_val, [])
                    is_sku_serial = sku_val in serial_skus

                    if not is_sku_serial:
                        # 1. PHÂN TÍCH SKU KHÔNG CÓ SERIAL (Đếm theo Số lượng)
                        erp_tot_qty = int(erp_qty_dict.get(sku_val, 0))
                        rem_erp_q = erp_tot_qty
                        for idx_ns, r in enumerate(s_rows):
                            sys_q = r["Số lượng"]
                            if rem_erp_q >= sys_q:
                                alloc_q = sys_q
                                rem_erp_q -= sys_q
                            else:
                                alloc_q = rem_erp_q
                                rem_erp_q = 0

                            if idx_ns == len(s_rows) - 1 and rem_erp_q > 0:
                                alloc_q += rem_erp_q
                                rem_erp_q = 0

                            diff_q = alloc_q - sys_q
                            r["Số Serial"] = "(Hàng không serial)"
                            r["Serial đã quét"] = f"{alloc_q} cái" if alloc_q > 0 else "0 cái"
                            r["Loại hàng"] = "Hàng không serial"
                            r["Đã kiểm"] = alloc_q
                            r["Dư/Thiếu"] = diff_q
                            if diff_q == 0:
                                r["Check đơn"] = f"Đã đếm đủ ({alloc_q}/{sys_q})"
                            elif diff_q < 0:
                                r["Check đơn"] = f"Bắn thiếu {abs(diff_q)} hàng ({alloc_q}/{sys_q})"
                            else:
                                r["Check đơn"] = f"Bắn thừa {diff_q} hàng ({alloc_q}/{sys_q})"
                            
                            r_clean = {k: v for k, v in r.items() if k != 'is_non_serial'}
                            detail_rows.append(r_clean)

                        if len(s_rows) == 0 and erp_tot_qty > 0:
                            detail_rows.append({
                                "Mã sản phẩm": sku_val, "Tên sản phẩm": "Sản phẩm không serial chưa có trong danh mục", "Part number": "", "ĐVT": "Cái",
                                "Số lượng": 0, "Số Serial": "(Hàng không serial)", "Serial đã quét": f"Đã đếm: {erp_tot_qty}", "Mã Bin": "",
                                "Tên Bin": "", "Loại hàng": "Không có tồn hệ thống",
                                "Đã kiểm": erp_tot_qty, "Dư/Thiếu": erp_tot_qty, "Check đơn": f"Bắn thừa hàng {erp_tot_qty}", "Note đơn hàng": ""
                            })

                    else:
                        # 2. PHÂN TÍCH SKU CÓ SERIAL (Quét theo Serial)
                        e_serials = list(erp_serials_by_sku.get(sku_val, []))
                        
                        # Ghép nối Serial khớp chính xác
                        for r in s_rows:
                            s_stock_ser = r["Số Serial"]
                            if s_stock_ser and s_stock_ser in e_serials:
                                r["Serial đã quét"] = s_stock_ser
                                r["Đã kiểm"] = 1
                                r["Dư/Thiếu"] = 0
                                r["Check đơn"] = "Đã quét đủ"
                                e_serials.remove(s_stock_ser)

                        # Các dòng chưa khớp -> Bắn thiếu
                        for r in s_rows:
                            if r["Đã kiểm"] == 0:
                                r["Serial đã quét"] = ""
                                r["Đã kiểm"] = 0
                                r["Dư/Thiếu"] = -1
                                r["Check đơn"] = "Bắn thiếu (Chưa quét)"

                            r_clean = {k: v for k, v in r.items() if k != 'is_non_serial'}
                            detail_rows.append(r_clean)

                        # Các Serial dư từ ERP
                        if len(e_serials) > 0:
                            has_stock = len(s_rows) > 0
                            first_name = s_rows[0]["Tên sản phẩm"] if has_stock else "Sản phẩm quét dư / chưa có trong danh mục"
                            first_uom = s_rows[0]["ĐVT"] if has_stock else "Cái"
                            for extra_ser in e_serials:
                                detail_rows.append({
                                    "Mã sản phẩm": sku_val, "Tên sản phẩm": first_name, "Part number": "", "ĐVT": first_uom,
                                    "Số lượng": 1, "Số Serial": "", "Serial đã quét": extra_ser, "Mã Bin": "",
                                    "Tên Bin": "", "Loại hàng": "Bắn dư serial" if has_stock else "Không có tồn hệ thống",
                                    "Đã kiểm": 1, "Dư/Thiếu": 1, "Check đơn": "Bắn dư serial", "Note đơn hàng": ""
                                })

                st.session_state.df_check_detail = pd.DataFrame(detail_rows)
                save_session_state(show_progress=True)
                st.success("Đã trích xuất & tạo thành công các bảng đối soát.")

# ==========================================
# TAB 2: KIỂM ĐẾM LẦN 2 (Chi tiết Serial theo Bin)
# ==========================================
with tabs[1]:
    st.subheader("Chi tiết danh sách Serial & ghi chú theo Bin")
    st.caption("Nhập hoặc quét trực tiếp vào ô 'Serial đã quét (ERP)' để bổ sung/thay thế. Xóa trắng ô Serial bắn dư để mở hộp thoại nhập lý do loại bỏ.")

    if st.session_state.pending_clear_serial:
        sku_p, ser_p = st.session_state.pending_clear_serial
        modal_clear_surplus(sku_p, ser_p)

    if st.session_state.df_check_detail is not None:
        df_dt_full = st.session_state.df_check_detail.copy()

        surplus_df = df_dt_full[df_dt_full['Check đơn'] == 'Bắn dư serial']
        if not surplus_df.empty:
            with st.expander("Bảng điều khiển nhanh: Xóa / Clear Serial Bắn dư khỏi Báo cáo", expanded=True):
                c_s1, c_s2 = st.columns([4, 2])
                with c_s1:
                    options_surplus = [f"{r['Mã sản phẩm']} - {r['Serial đã quét']} ({r['Tên sản phẩm']})" for _, r in surplus_df.iterrows()]
                    selected_surplus_item = st.selectbox("Chọn Serial dư cần loại bỏ:", options_surplus, key="sel_surplus_item")
                with c_s2:
                    st.write("")
                    st.write("")
                    if st.button("Loại bỏ Serial đã chọn", type="primary", use_container_width=True):
                        parts = selected_surplus_item.split(" - ")
                        sku_sel = parts[0].strip()
                        ser_sel = parts[1].split(" (")[0].strip()
                        st.session_state.pending_clear_serial = (sku_sel, ser_sel)
                        st.rerun()

        cf1, cf2 = st.columns([3, 3])
        with cf1:
            status_dt = st.selectbox(
                "Lọc theo Trạng thái Serial:",
                ["Tất cả Serial", "Chỉ Serial chênh lệch (Dư / Thiếu)", "Thiếu (Chưa quét)", "Bắn sai serial", "Bắn dư serial", "Đã loại bỏ Serial dư"],
                index=0,
                key="status_dt"
            )
        with cf2:
            search_dt = st.text_input("Tìm kiếm SKU / Serial / Bin:", "", key="search_dt")

        df_dt_filtered = df_dt_full.copy()
        if status_dt.startswith("Chỉ Serial"):
            df_dt_filtered = df_dt_filtered[df_dt_filtered['Dư/Thiếu'] != 0]
        elif status_dt == "Thiếu (Chưa quét)":
            df_dt_filtered = df_dt_filtered[df_dt_filtered['Dư/Thiếu'] < 0]
        elif status_dt == "Bắn sai serial":
            df_dt_filtered = df_dt_filtered[df_dt_filtered['Check đơn'] == "Bắn sai serial"]
        elif status_dt == "Bắn dư serial":
            df_dt_filtered = df_dt_filtered[df_dt_filtered['Check đơn'] == "Bắn dư serial"]
        elif status_dt == "Đã loại bỏ Serial dư":
            df_dt_filtered = df_dt_filtered[df_dt_filtered['Check đơn'] == "Đã loại bỏ Serial dư"]

        if search_dt.strip():
            kw_d = search_dt.strip().lower()
            df_dt_filtered = df_dt_filtered[
                df_dt_filtered['Mã sản phẩm'].astype(str).str.lower().str.contains(kw_d) |
                df_dt_filtered['Số Serial'].astype(str).str.lower().str.contains(kw_d) |
                df_dt_filtered['Serial đã quét'].astype(str).str.lower().str.contains(kw_d) |
                df_dt_filtered['Mã Bin'].astype(str).str.lower().str.contains(kw_d)
            ]

        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Tổng dòng Serial", len(df_dt_filtered))
        s2.metric("Số Serial Đã kiểm (Quét đủ)", len(df_dt_filtered[df_dt_filtered['Đã kiểm'] == 1]))
        s3.metric("Số Serial Thiếu", len(df_dt_filtered[df_dt_filtered['Dư/Thiếu'] < 0]))
        s4.metric("Số Serial Dư / Sai mã", len(df_dt_filtered[df_dt_filtered['Dư/Thiếu'] > 0]))

        edited_dt = st.data_editor(
            df_dt_filtered,
            column_config={
                "Mã sản phẩm": st.column_config.TextColumn("Mã sản phẩm", disabled=True),
                "Tên sản phẩm": st.column_config.TextColumn("Tên sản phẩm", disabled=True),
                "Part number": st.column_config.TextColumn("Part number", disabled=True),
                "ĐVT": st.column_config.TextColumn("ĐVT", disabled=True),
                "Số lượng": st.column_config.NumberColumn("Số lượng", disabled=True),
                "Số Serial": st.column_config.TextColumn("Số Serial (Sổ sách)", disabled=True),
                "Serial đã quét": st.column_config.TextColumn("Serial đã quét (ERP - Nhập/Bắn bù)", disabled=False),
                "Mã Bin": st.column_config.TextColumn("Mã Bin", disabled=True),
                "Tên Bin": st.column_config.TextColumn("Tên Bin / Khu vực", disabled=True),
                "Loại hàng": st.column_config.TextColumn("Loại hàng", disabled=True),
                "Đã kiểm": st.column_config.NumberColumn("Đã kiểm (1/0)", disabled=True),
                "Dư/Thiếu": st.column_config.NumberColumn("Dư/Thiếu (1/-1)", disabled=True),
                "Check đơn": st.column_config.TextColumn("Check đơn", disabled=True),
                "Note đơn hàng": st.column_config.TextColumn("Note đơn hàng (Lịch sử & Ghi chú)")
            },
            hide_index=True,
            use_container_width=True,
            key="editor_check_detail"
        )

        has_changes = False
        for idx, r in edited_dt.iterrows():
            sku_val = r['Mã sản phẩm']
            s_stock = str(r['Số Serial']).strip()
            item_type = str(r['Loại hàng']).strip()
            user_note = str(r['Note đơn hàng']).strip() if pd.notna(r['Note đơn hàng']) else ""

            if item_type == "Hàng không serial" or s_stock == "(Hàng không serial)":
                bin_m = str(r['Mã Bin']).strip() if pd.notna(r['Mã Bin']) else ""
                mask = (st.session_state.df_check_detail['Mã sản phẩm'] == sku_val) & (st.session_state.df_check_detail['Mã Bin'] == bin_m)
                st.session_state.df_check_detail.loc[mask, 'Note đơn hàng'] = user_note
                continue

            new_scanned = str(r['Serial đã quét']).strip() if pd.notna(r['Serial đã quét']) else ""

            if s_stock:
                mask = (st.session_state.df_check_detail['Số Serial'] == s_stock) & (st.session_state.df_check_detail['Mã sản phẩm'] == sku_val)
            else:
                mask = (st.session_state.df_check_detail['Mã sản phẩm'] == sku_val) & (st.session_state.df_check_detail['Serial đã quét'] == new_scanned)

            old_row = st.session_state.df_check_detail.loc[mask]
            if not old_row.empty:
                old_scanned = str(old_row.iloc[0]['Serial đã quét']).strip() if pd.notna(old_row.iloc[0]['Serial đã quét']) else ""
                old_checked = old_row.iloc[0]['Đã kiểm']
                old_chk_reason = old_row.iloc[0]['Check đơn']

                if new_scanned != old_scanned:
                    has_changes = True
                    if new_scanned == "" and (old_chk_reason == "Bắn dư serial" or old_scanned != ""):
                        st.session_state.pending_clear_serial = (sku_val, old_scanned)
                        st.rerun()

                    elif new_scanned != "":
                        if old_checked == 0 or old_scanned == "":
                            st.session_state.df_check_detail.loc[mask, 'Serial đã quét'] = new_scanned
                            st.session_state.df_check_detail.loc[mask, 'Đã kiểm'] = 1
                            st.session_state.df_check_detail.loc[mask, 'Dư/Thiếu'] = 0
                            st.session_state.df_check_detail.loc[mask, 'Check đơn'] = "Đã kiểm đếm bổ sung Lần 2"
                            history_tag = f"[Bổ sung Serial Lần 2: {new_scanned}]"
                            combined_n = f"{user_note} {history_tag}".strip()
                            st.session_state.df_check_detail.loc[mask, 'Note đơn hàng'] = combined_n
                            l2_mask = st.session_state.df_count_l2['Mã vật tư'] == sku_val
                            st.session_state.df_count_l2.loc[l2_mask, 'Số liệu thực tế đếm lại lần 2'] += 1
                        else:
                            st.session_state.df_check_detail.loc[mask, 'Serial đã quét'] = new_scanned
                            st.session_state.df_check_detail.loc[mask, 'Đã kiểm'] = 1
                            st.session_state.df_check_detail.loc[mask, 'Dư/Thiếu'] = 0
                            st.session_state.df_check_detail.loc[mask, 'Check đơn'] = "Đã sửa Serial (Lần 2)"
                            history_tag = f"[Sửa Serial: {old_scanned} -> {new_scanned}]"
                            combined_n = f"{user_note} {history_tag}".strip()
                            st.session_state.df_check_detail.loc[mask, 'Note đơn hàng'] = combined_n
                else:
                    st.session_state.df_check_detail.loc[mask, 'Note đơn hàng'] = user_note

        if has_changes:
            save_session_state(show_progress=True)

    else:
        st.warning("Vui lòng tải lên file dữ liệu tại Tab 1 trước.")

# ==========================================
# (Nội dung Kiểm đếm lần 2 được xử lý ngầm, không hiển thị thành tab riêng)


# ==========================================
# TAB 3: KẾT QUẢ KIỂM LẦN 2 (Bảng chỉnh số lượng bù)
# ==========================================
with tabs[2]:
    st.subheader("Kết quả kiểm lần 2 – Bảng xử lý số liệu chênh lệch")
    st.caption("Hiển thị TẤT CẢ SKU. Cập nhật số lượng vào cột 'Số liệu thực tế đếm lại lần 2' để bù chênh lệch.")

    if st.session_state.df_count_l2 is not None:
        df_l2_full = st.session_state.df_count_l2.copy()

        c_filter1, c_filter2, c_filter3 = st.columns([2, 2, 2])
        with c_filter1:
            view_mode_l2 = st.radio(
                "Chế độ hiển thị:",
                ["Chỉ SKU chênh lệch (Cần xử lý)", "Tất cả SKU"],
                index=0,
                key="view_mode_l2"
            )
        with c_filter2:
            status_l2 = st.selectbox(
                "Trạng thái đếm Lần 2:",
                ["Tất cả", "Lệch chưa đếm bù", "Đã nhập đếm bù Lần 2", "Đã khớp sau đếm bù"],
                key="status_l2"
            )
        with c_filter3:
            search_l2 = st.text_input("Tìm kiếm Mã / Tên vật tư:", "", key="search_l2")

        df_l2_filtered = df_l2_full.copy()

        if view_mode_l2.startswith("Chỉ SKU"):
            df_l2_filtered = df_l2_filtered[
                (df_l2_filtered['Số liệu thực tế đã đếm lần 1'] != df_l2_filtered['Số lượng sổ sách (hệ thống)']) |
                (df_l2_filtered['Số liệu thực tế đếm lại lần 2'] > 0) |
                (df_l2_filtered['Chênh lệch sau kiểm đếm lần 2'] != 0)
            ]

        if status_l2 == "Lệch chưa đếm bù":
            df_l2_filtered = df_l2_filtered[(df_l2_filtered['Chênh lệch sau kiểm đếm lần 2'] != 0) & (df_l2_filtered['Số liệu thực tế đếm lại lần 2'] == 0)]
        elif status_l2 == "Đã nhập đếm bù Lần 2":
            df_l2_filtered = df_l2_filtered[df_l2_filtered['Số liệu thực tế đếm lại lần 2'] > 0]
        elif status_l2 == "Đã khớp sau đếm bù":
            df_l2_filtered = df_l2_filtered[df_l2_filtered['Chênh lệch sau kiểm đếm lần 2'] == 0]

        if search_l2.strip():
            kw = search_l2.strip().lower()
            df_l2_filtered = df_l2_filtered[
                df_l2_filtered['Mã vật tư'].astype(str).str.lower().str.contains(kw) |
                df_l2_filtered['Tên vật tư'].astype(str).str.lower().str.contains(kw)
            ]

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("Tổng SKU", len(df_l2_filtered))
        m2.metric("Tổng Sổ sách (A)", f"{int(df_l2_filtered['Số lượng sổ sách (hệ thống)'].sum()):,}")
        m3.metric("Tổng Lần 1 (B)", f"{int(df_l2_filtered['Số liệu thực tế đã đếm lần 1'].sum()):,}")
        m4.metric("Tổng Lần 2 (C)", f"{int(df_l2_filtered['Số liệu thực tế đếm lại lần 2'].sum()):,}")
        m5.metric("Tổng Sau Cùng (B+C)", f"{int(df_l2_filtered['Kết quả xử lý sau khi đếm lại lần 2'].sum()):,}")
        diff_total_l2 = int(df_l2_filtered['Chênh lệch sau kiểm đếm lần 2'].sum())
        m6.metric("Tổng Chênh Lệch", f"{diff_total_l2:,}", delta=diff_total_l2, delta_color="inverse")

        edited_l2 = st.data_editor(
            df_l2_filtered,
            column_config={
                "Mã vật tư": st.column_config.TextColumn("Mã vật tư", disabled=True),
                "Tên vật tư": st.column_config.TextColumn("Tên vật tư", disabled=True),
                "Đvt": st.column_config.TextColumn("Đvt", disabled=True),
                "Số lượng sổ sách (hệ thống)": st.column_config.NumberColumn("Số lượng sổ sách (A)", disabled=True),
                "Số liệu thực tế đã đếm lần 1": st.column_config.NumberColumn("Thực tế Lần 1 (B)", disabled=True),
                "Số liệu thực tế đếm lại lần 2": st.column_config.NumberColumn("Đếm lại Lần 2 (C)", min_value=0, step=1, disabled=False),
                "Chênh lệch sau kiểm đếm lần 2": st.column_config.NumberColumn("Chênh lệch Lần 2", disabled=True),
                "Kết quả xử lý sau khi đếm lại lần 2": st.column_config.NumberColumn("Tổng thực tế (B+C)", disabled=True),
                "Note mã đơn": st.column_config.TextColumn("Note mã đơn xuất / Lý do", disabled=False)
            },
            hide_index=True,
            use_container_width=True,
            key="editor_count_l2"
        )

        has_l2_changes = False
        for idx, row in edited_l2.iterrows():
            sku_val = row['Mã vật tư']
            l2_val = row['Số liệu thực tế đếm lại lần 2']
            note_val = str(row['Note mã đơn']).strip() if pd.notna(row['Note mã đơn']) else ""
            mask = st.session_state.df_count_l2['Mã vật tư'] == sku_val
            if len(st.session_state.df_count_l2.loc[mask]) == 0:
                continue
            old_l2 = st.session_state.df_count_l2.loc[mask, 'Số liệu thực tế đếm lại lần 2'].values[0]
            old_note = str(st.session_state.df_count_l2.loc[mask, 'Note mã đơn'].values[0]).strip()
            if l2_val != old_l2 or note_val != old_note:
                has_l2_changes = True
                st.session_state.df_count_l2.loc[mask, 'Số liệu thực tế đếm lại lần 2'] = l2_val
                st.session_state.df_count_l2.loc[mask, 'Chênh lệch sau kiểm đếm lần 2'] = (st.session_state.df_count_l2.loc[mask, 'Số liệu thực tế đã đếm lần 1'] + l2_val) - st.session_state.df_count_l2.loc[mask, 'Số lượng sổ sách (hệ thống)']
                st.session_state.df_count_l2.loc[mask, 'Kết quả xử lý sau khi đếm lại lần 2'] = st.session_state.df_count_l2.loc[mask, 'Số liệu thực tế đã đếm lần 1'] + l2_val
                st.session_state.df_count_l2.loc[mask, 'Note mã đơn'] = note_val
        if has_l2_changes:
            save_session_state(show_progress=True)
    else:
        st.warning("Vui lòng tải lên file dữ liệu tại Tab 1 trước.")

# ==========================================
# TAB 4: BẢNG TỔNG HỢP CHÊnh lệch
# ==========================================
with tabs[3]:
    st.subheader("Bảng tổng hợp chênh lệch & nguyên nhân sai lệch")
    st.caption("Tổng hợp kết quả đối soát tự động theo Mã sản phẩm (SKU) và tự động nhận diện lý do sai lệch")
    
    if st.session_state.df_count_l2 is not None and st.session_state.df_check_detail is not None:
        df_l2_c = st.session_state.df_count_l2.copy()
        df_dt_c = st.session_state.df_check_detail.copy()
        
        detail_notes = df_dt_c.groupby('Mã sản phẩm')['Note đơn hàng'].apply(
            lambda s: " | ".join(list(dict.fromkeys([str(v).strip() for v in s if pd.notna(v) and str(v).strip() != ''])))
        ).to_dict()

        wrong_serial_skus = set(df_dt_c[df_dt_c['Check đơn'] == 'Bắn sai serial']['Mã sản phẩm'])
        surplus_skus = set(df_dt_c[df_dt_c['Check đơn'] == 'Bắn dư serial']['Mã sản phẩm'])
        missing_skus = set(df_dt_c[df_dt_c['Check đơn'] == 'Bắn thiếu (Chưa quét)']['Mã sản phẩm'])

        main_rows = []
        for idx, row in df_l2_c.iterrows():
            sku = str(row['Mã vật tư']).strip()
            name = str(row['Tên vật tư']).strip()
            qty_sys = int(row['Số lượng sổ sách (hệ thống)'])
            qty_l1 = int(row['Số liệu thực tế đã đếm lần 1'])
            qty_l2 = int(row['Số liệu thực tế đếm lại lần 2'])
            total_act = qty_l1 + qty_l2
            diff = total_act - qty_sys
            
            l2_note = str(row['Note mã đơn']).strip() if pd.notna(row['Note mã đơn']) else ""
            dt_note = detail_notes.get(sku, "").strip()
            
            combined_note_list = list(dict.fromkeys([n for n in [dt_note, l2_note] if n]))
            combined_note = " | ".join(combined_note_list)

            if sku in wrong_serial_skus:
                chk_status = "Bắn sai serial"
            elif sku in surplus_skus:
                chk_status = "Bắn dư serial"
            elif qty_sys > qty_l1 and total_act == qty_sys:
                chk_status = "Bắn thiếu hàng - Kiểm lần 2 đủ"
            elif qty_sys > qty_l1 and total_act < qty_sys:
                chk_status = f"Bắn thiếu hàng - Kiểm lần 2 thiếu {abs(diff)}"
            elif total_act < qty_sys:
                chk_status = "Bắn thiếu hàng"
            elif total_act > qty_sys:
                chk_status = f"Bắn thừa hàng {diff}"
            else:
                chk_status = "Khớp hoàn toàn"

            main_rows.append({
                "SKU": sku,
                "Tên hàng hóa": name,
                "Số lượng sổ sách (hệ thống)": qty_sys,
                "Số lượng thực tế": total_act,
                "Chênh lệch": diff,
                "Check": chk_status,
                "Note mã đơn": combined_note
            })

        df_main_full = pd.DataFrame(main_rows)
        
        c_mfilter1, c_mfilter2, c_mfilter3 = st.columns([2, 2, 2])
        with c_mfilter1:
            view_mode_main = st.radio(
                "Chế độ hiển thị:",
                ["Chỉ SKU chênh lệch / Sai lý do (Cần xử lý)", "Tất cả SKU"],
                index=0,
                key="view_mode_main"
            )
        with c_mfilter2:
            status_main = st.selectbox(
                "Lọc theo Phân loại Check:",
                ["Tất cả", "Bắn sai serial", "Bắn dư serial", "Bắn thiếu hàng", "Bắn thừa hàng", "Khớp hoàn toàn", "Đã loại bỏ Serial dư"],
                key="status_main"
            )
        with c_mfilter3:
            search_main = st.text_input("Tìm kiếm SKU / Tên hàng hóa:", "", key="search_main")

        df_main_filtered = df_main_full.copy()

        if view_mode_main.startswith("Chỉ SKU"):
            df_main_filtered = df_main_filtered[df_main_filtered['Check'] != "Khớp hoàn toàn"]

        if status_main == "Bắn sai serial":
            df_main_filtered = df_main_filtered[df_main_filtered['Check'] == "Bắn sai serial"]
        elif status_main == "Bắn dư serial":
            df_main_filtered = df_main_filtered[df_main_filtered['Check'] == "Bắn dư serial"]
        elif status_main == "Bắn thiếu hàng":
            df_main_filtered = df_main_filtered[df_main_filtered['Check'].str.contains("Bắn thiếu", na=False)]
        elif status_main == "Bắn thừa hàng":
            df_main_filtered = df_main_filtered[df_main_filtered['Check'].str.contains("Bắn thừa", na=False)]
        elif status_main == "Khớp hoàn toàn":
            df_main_filtered = df_main_filtered[df_main_filtered['Check'] == "Khớp hoàn toàn"]
        elif status_main == "Đã loại bỏ Serial dư":
            cleared_skus = set(st.session_state.df_check_detail[st.session_state.df_check_detail['Check đơn'] == 'Đã loại bỏ Serial dư']['Mã sản phẩm'])
            df_main_filtered = df_main_filtered[df_main_filtered['SKU'].isin(cleared_skus)]

        if search_main.strip():
            kw_m = search_main.strip().lower()
            df_main_filtered = df_main_filtered[
                df_main_filtered['SKU'].astype(str).str.lower().str.contains(kw_m) |
                df_main_filtered['Tên hàng hóa'].astype(str).str.lower().str.contains(kw_m)
            ]

        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("Số lượng SKU", len(df_main_filtered))
        k2.metric("Tổng SL Sổ sách", f"{int(df_main_filtered['Số lượng sổ sách (hệ thống)'].sum()):,}")
        k3.metric("Tổng SL Thực tế", f"{int(df_main_filtered['Số lượng thực tế'].sum()):,}")
        diff_tot_main = int(df_main_filtered['Chênh lệch'].sum())
        k4.metric("Tổng SL Chênh lệch", f"{diff_tot_main:,}", delta=diff_tot_main, delta_color="inverse")
        k5.metric("Số SKU Lỗi / Dư Serial", len(df_main_filtered[df_main_filtered['Check'].str.contains("Bắn sai|Bắn dư", na=False)]))

        def highlight_check(val):
            v = str(val)
            if "Bắn sai serial" in v:
                return 'background-color: #ffff99; font-weight: bold; color: #806000;'
            elif "Bắn dư serial" in v:
                return 'background-color: #ffe699; font-weight: bold; color: #804000;'
            elif "Bắn thiếu" in v:
                return 'background-color: #ffcccc; color: darkred;'
            elif "Khớp" in v:
                return 'background-color: #ccffcc; color: darkgreen;'
            return ''

        st.dataframe(
            df_main_filtered.style.map(highlight_check, subset=['Check']),
            use_container_width=True,
            hide_index=True,
            column_config={
                "SKU": st.column_config.TextColumn("SKU", width=120),
                "Tên hàng hóa": st.column_config.TextColumn("Tên hàng hóa", width=280),
                "Số lượng sổ sách (hệ thống)": st.column_config.NumberColumn("Sổ sách", width=90),
                "Số lượng thực tế": st.column_config.NumberColumn("Thực tế", width=90),
                "Chênh lệch": st.column_config.NumberColumn("Chênh lệch", width=90),
                "Check": st.column_config.TextColumn("Kết quả kiểm", width=220),
                "Note mã đơn": st.column_config.TextColumn("Note / Ghi chú", width=400),
            }
        )
    else:
        st.warning("Vui lòng tải dữ liệu tại Tab 1 trước.")

# ==========================================
# TAB 4: XUẤT BÁO CÁO EXCEL & PDF
# ==========================================
with tabs[4]:
    st.subheader("Xuất biên bản & báo cáo kiểm kê")
    
    if st.session_state.df_count_l2 is not None and st.session_state.df_check_detail is not None:
        
        with st.container(border=True):
            st.markdown("**Thông tin Biên bản Kiểm kê Hàng tồn kho**")
            col_meta1, col_meta2 = st.columns(2)
            with col_meta1:
                report_date = st.date_input("Ngày kiểm kê:", datetime.date.today())
                location_name = st.text_input("Tên Kho / Chi nhánh:", "CP74 - KHÁNH HỘI | Địa chỉ: 162 - 164 Khánh Hội, Phường Khánh Hội, TPHCM")
            with col_meta2:
                mgr_name = st.text_input("Quản lý Đơn vị:", "Nguyễn Tuấn Vũ")
                staff_name = st.text_input("Đại diện Kiểm kê / Kho:", "Nguyễn Xuân Lộc")

        st.markdown("---")

        col_exp1, col_exp2 = st.columns(2)
        
        df_l2_c = st.session_state.df_count_l2
        df_dt_c = st.session_state.df_check_detail
        
        wrong_skus_dict = df_dt_c[df_dt_c['Check đơn'].str.contains('Bắn sai|Sửa Serial', na=False)].groupby('Mã sản phẩm')['Serial đã quét'].apply(
            lambda s: ", ".join([str(v).strip() for v in s if pd.notna(v) and str(v).strip() != ''])
        ).to_dict()

        dt_notes_c = df_dt_c.groupby('Mã sản phẩm')['Note đơn hàng'].apply(
            lambda s: " ".join([clean_export_note(v) for v in s if pd.notna(v) and clean_export_note(v) != ''])
        ).to_dict()

        main_out = []
        for idx_r, r in df_l2_c.iterrows():
            sk = str(r['Mã vật tư']).strip()
            nm = str(r['Tên vật tư']).strip()
            uom = str(r['Đvt']).strip()
            q_sys = int(r['Số lượng sổ sách (hệ thống)'])
            q_l1 = int(r['Số liệu thực tế đã đếm lần 1'])
            q_l2 = int(r['Số liệu thực tế đếm lại lần 2'])
            tot = q_l1 + q_l2
            df_val = tot - q_sys
            
            l2_n = clean_export_note(r['Note mã đơn'])
            dt_n = dt_notes_c.get(sk, "").strip()
            w_ser = wrong_skus_dict.get(sk, "").strip()

            note_parts = [n for n in [w_ser, dt_n, l2_n] if n]
            comb_n = " | ".join(list(dict.fromkeys(note_parts)))

            if sk in wrong_skus_dict:
                chk = "Bắn sai serial"
            elif q_sys > q_l1 and tot == q_sys:
                chk = "Bắn thiếu hàng - Kiểm lần 2 đủ"
            elif q_sys > q_l1 and tot < q_sys:
                chk = f"Bắn thiếu hàng - Kiểm lần 2 thiếu {abs(df_val)}"
            elif tot < q_sys:
                chk = "Bắn thiếu hàng"
            elif tot > q_sys:
                chk = f"Bắn thừa hàng {df_val}"
            else:
                chk = "Khớp hoàn toàn"

            main_out.append({
                "STT": idx_r + 1,
                "SKU": sk,
                "Tên hàng hóa": nm,
                "Đơn vị tính": uom,
                "Số lượng sổ sách (hệ thống)": q_sys,
                "Số lượng thực tế": tot,
                "Chênh lệch": df_val,
                "Check": chk,
                "Ghi chú/ Tình trạng hàng hóa": comb_n
            })

        df_th_out = pd.DataFrame(main_out)
        sum_sys = df_th_out['Số lượng sổ sách (hệ thống)'].sum()
        sum_act = df_th_out['Số lượng thực tế'].sum()
        sum_diff = df_th_out['Chênh lệch'].sum()

        output_excel = io.BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            wb = writer.book
            ws = wb.create_sheet(title="TH-HANG HOA")
            
            if os.path.exists('phongvu_logo.png'):
                try:
                    logo_img = OpenpyxlImage('phongvu_logo.png')
                    logo_img.width = 140
                    logo_img.height = 35
                    ws.add_image(logo_img, 'A1')
                except Exception:
                    pass
            else:
                ws['A1'] = "PHONGVU.VN"

            font_title = Font(name='Arial', size=14, bold=True)
            font_subtitle = Font(name='Arial', size=11, bold=True, color='CC0000')
            font_bold = Font(name='Arial', size=10, bold=True)
            font_normal = Font(name='Arial', size=9)
            font_red = Font(name='Arial', size=9, color='CC0000', bold=True)
            font_hdr = Font(name='Arial', size=10, bold=True)
            fill_hdr = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

            thin_border = Border(left=Side(style='thin', color='AAAAAA'), right=Side(style='thin', color='AAAAAA'), top=Side(style='thin', color='AAAAAA'), bottom=Side(style='thin', color='AAAAAA'))
            double_bottom_border = Border(left=Side(style='thin', color='AAAAAA'), right=Side(style='thin', color='AAAAAA'), top=Side(style='thin', color='AAAAAA'), bottom=Side(style='double', color='000000'))

            ws['A2'] = "CÔNG TY CỔ PHẦN THƯƠNG MẠI DỊCH VỤ PHONG VŨ"; ws['A2'].font = font_normal
            ws['C4'] = "BIÊN BẢN KIỂM KÊ HÀNG TỒN KHO"; ws['C4'].font = font_title
            ws['C5'] = f"NGÀY {report_date.strftime('%d.%m.%Y')}"; ws['C5'].font = font_subtitle
            ws['A7'] = f"Hôm nay, vào lúc __ giờ __ phút, ngày {report_date.strftime('%d')} tháng {report_date.strftime('%m')} năm {report_date.strftime('%Y')}"
            ws['A8'] = f"Tại kho: {location_name} , chúng tôi gồm có:"
            ws['A9'] = f"Họ tên : {mgr_name}          Chức danh: Quản lý Đơn vị"
            ws['A10'] = f"Họ tên : {staff_name}          Chức danh: Đại diện Kiểm kê / Kho"
            
            ws['E11'] = sum_sys; ws['E11'].font = font_bold; ws['E11'].alignment = Alignment(horizontal='right')
            ws['F11'] = sum_act; ws['F11'].font = font_bold; ws['F11'].alignment = Alignment(horizontal='right')
            ws['G11'] = f"({abs(sum_diff)})" if sum_diff < 0 else str(sum_diff); ws['G11'].font = font_red; ws['G11'].alignment = Alignment(horizontal='right')

            headers_excel = ['STT', 'SKU', 'Tên hàng hóa', 'Đơn vị tính', 'Số lượng sổ sách (hệ thống)', 'Số lượng thực tế', 'Chênh lệch', 'Ghi chú/ Tình trạng hàng hóa']
            for col_idx, h_text in enumerate(headers_excel, 1):
                cell = ws.cell(row=13, column=col_idx, value=h_text)
                cell.font = font_hdr; cell.fill = fill_hdr; cell.alignment = Alignment(horizontal='center' if col_idx in [1, 4] else ('right' if col_idx in [5, 6, 7] else 'left'), vertical='center', wrap_text=True); cell.border = thin_border
            ws.row_dimensions[13].height = 28

            current_row = 14
            for _, r_data in df_th_out.iterrows():
                diff_val = r_data['Chênh lệch']
                diff_disp = f"({abs(diff_val)})" if diff_val < 0 else ("-" if diff_val == 0 else str(diff_val))
                row_vals = [r_data['STT'], str(r_data['SKU']), r_data['Tên hàng hóa'], r_data['Đơn vị tính'], r_data['Số lượng sổ sách (hệ thống)'], r_data['Số lượng thực tế'], diff_disp, clean_export_note(r_data['Ghi chú/ Tình trạng hàng hóa'])]
                for col_idx, val in enumerate(row_vals, 1):
                    c = ws.cell(row=current_row, column=col_idx, value=val)
                    c.font = font_red if (col_idx == 7 and diff_val != 0) else font_normal
                    c.alignment = Alignment(horizontal='center' if col_idx in [1, 4] else ('right' if col_idx in [5, 6, 7] else 'left'), vertical='center'); c.border = thin_border
                current_row += 1

            ws.cell(row=current_row, column=1, value="Tổng cộng").font = font_red; ws.cell(row=current_row, column=5, value=sum_sys).font = font_red; ws.cell(row=current_row, column=6, value=sum_act).font = font_red; ws.cell(row=current_row, column=7, value=f"({abs(sum_diff)})" if sum_diff < 0 else str(sum_diff)).font = font_red
            for col_idx in range(1, 9): ws.cell(row=current_row, column=col_idx).border = double_bottom_border
            
            column_widths = {'A': 8, 'B': 18, 'C': 55, 'D': 14, 'E': 24, 'F': 20, 'G': 16, 'H': 40}
            for col_letter, width in column_widths.items(): ws.column_dimensions[col_letter].width = width

            st.session_state.df_count_l2.to_excel(writer, sheet_name="Kiểm đếm lần 2", index=False)
            df_th_out[['SKU', 'Tên hàng hóa', 'Số lượng sổ sách (hệ thống)', 'Số lượng thực tế', 'Chênh lệch', 'Check', 'Ghi chú/ Tình trạng hàng hóa']].to_excel(writer, sheet_name="main", index=False)
            st.session_state.df_check_detail.to_excel(writer, sheet_name="check", index=False)

            # ---- Định dạng các sheet phụ (Kiểm đếm lần 2, main, check) ----
            sheet_col_widths = {
                "Kiểm đếm lần 2": {
                    "Mã vật tư": 18, "Tên vật tư": 45, "Đvt": 10,
                    "Số lượng sổ sách (hệ thống)": 18, "Số liệu thực tế đã đếm lần 1": 18,
                    "Số liệu thực tế đếm lại lần 2": 18, "Chênh lệch sau kiểm đếm lần 2": 18,
                    "Kết quả xử lý sau khi đếm lại lần 2": 22, "Note mã đơn": 50,
                },
                "main": {
                    "SKU": 18, "Tên hàng hóa": 45, "Số lượng sổ sách (hệ thống)": 18,
                    "Số lượng thực tế": 18, "Chênh lệch": 14, "Check": 30,
                    "Ghi chú/ Tình trạng hàng hóa": 55,
                },
                "check": {
                    "Mã sản phẩm": 18, "Tên sản phẩm": 40, "Part number": 20, "ĐVT": 10,
                    "Số lượng": 12, "Số Serial": 30, "Serial đã quét": 30,
                    "Mã Bin": 16, "Tên Bin": 16, "Loại hàng": 20,
                    "Đã kiểm": 12, "Dư/Thiếu": 12, "Check đơn": 35, "Note đơn hàng": 55,
                },
            }

            fill_hdr_sub = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            font_hdr_sub = Font(name='Arial', size=10, bold=True)
            font_data_sub = Font(name='Arial', size=9)
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
            border_sub = Border(
                left=Side(style='thin', color='BBBBBB'), right=Side(style='thin', color='BBBBBB'),
                top=Side(style='thin', color='BBBBBB'), bottom=Side(style='thin', color='BBBBBB')
            )

            for s_name in wb.sheetnames:
                if s_name == "TH-HANG HOA":
                    continue
                ws_sub = wb[s_name]
                widths = sheet_col_widths.get(s_name, {})

                # Style header row
                for cell in ws_sub[1]:
                    cell.font = font_hdr_sub
                    cell.fill = fill_hdr_sub
                    cell.alignment = align_center
                    cell.border = border_sub
                ws_sub.row_dimensions[1].height = 30
                ws_sub.freeze_panes = 'A2'

                # Style data rows
                for row in ws_sub.iter_rows(min_row=2, max_row=ws_sub.max_row):
                    for cell in row:
                        cell.font = font_data_sub
                        cell.alignment = align_left
                        cell.border = border_sub

                # Set column widths
                for col in ws_sub.columns:
                    col_letter = get_column_letter(col[0].column)
                    header_val = str(col[0].value or '')
                    if header_val in widths:
                        ws_sub.column_dimensions[col_letter].width = widths[header_val]
                    else:
                        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                        ws_sub.column_dimensions[col_letter].width = min(max_len + 3, 50)

        excel_bytes = output_excel.getvalue()

        with col_exp1:
            with st.container(border=True):
                st.markdown("**1. File Báo cáo Excel Biên bản**")
                st.download_button(label="Tải xuống File Excel Biên Bản (.xlsx)", data=excel_bytes, file_name=f"Bien_Ban_Kiem_Ke_Hang_Ton_Kho_{report_date.strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)

        with col_exp2:
            with st.container(border=True):
                st.markdown("**2. File Báo cáo PDF Biên bản**")
                pdf_output = io.BytesIO()
                doc_pdf = SimpleDocTemplate(pdf_output, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
                story_pdf = []
                
                f_name = 'ArialVN' if os.path.exists(pdf_font_path) else 'Helvetica'
                style_head = ParagraphStyle('Head', fontName=f_name, fontSize=12, leading=15, textColor=colors.HexColor('#003399'))
                style_title = ParagraphStyle('Title', fontName=f_name, fontSize=16, leading=20, alignment=1, textColor=colors.HexColor('#000000'))
                style_date_sub = ParagraphStyle('DateSub', fontName=f_name, fontSize=12, leading=16, alignment=1, textColor=colors.HexColor('#CC0000'))
                style_meta = ParagraphStyle('Meta', fontName=f_name, fontSize=9, leading=12)
                style_tbl_hdr = ParagraphStyle('TblHdr', fontName=f_name, fontSize=9, leading=11, alignment=1, textColor=colors.HexColor('#000000'))
                style_tbl_cell = ParagraphStyle('TblCell', fontName=f_name, fontSize=8, leading=10)
                style_tbl_cell_red = ParagraphStyle('TblCellRed', fontName=f_name, fontSize=8, leading=10, textColor=colors.HexColor('#CC0000'))
                
                # Thay thế chữ PHONGVU.VN bằng ảnh Logo trực tiếp tại vị trí góc trên trái
                if os.path.exists('phongvu_logo.png'):
                    try:
                        logo_img_pdf = ReportlabImage('phongvu_logo.png', width=130, height=32)
                        logo_img_pdf.hAlign = 'LEFT'
                        story_pdf.append(logo_img_pdf)
                        story_pdf.append(Spacer(1, 4))
                    except Exception:
                        story_pdf.append(Paragraph("<b>PHONGVU.VN</b>", style_head))
                else:
                    story_pdf.append(Paragraph("<b>PHONGVU.VN</b>", style_head))

                story_pdf.append(Paragraph("<b>CÔNG TY CỔ PHẦN THƯƠNG MẠI DỊCH VỤ PHONG VŨ</b>", style_meta))
                story_pdf.append(Spacer(1, 10))
                story_pdf.append(Paragraph("<b>BIÊN BẢN KIỂM KÊ HÀNG TỒN KHO</b>", style_title))
                story_pdf.append(Paragraph(f"<b>NGÀY {report_date.strftime('%d.%m.%Y')}</b>", style_date_sub))
                story_pdf.append(Spacer(1, 10))
                
                style_sig_title = ParagraphStyle('SigTitle', fontName=f_name, fontSize=9, leading=12, alignment=1)
                style_sig_sub = ParagraphStyle('SigSub', fontName=f_name, fontSize=8, leading=10, alignment=1, textColor=colors.HexColor('#555555'))

                # Metadata
                story_pdf.append(Paragraph(f"Hôm nay, vào lúc __ giờ __ phút, ngày {report_date.strftime('%d')} tháng {report_date.strftime('%m')} năm {report_date.strftime('%Y')}", style_meta))
                story_pdf.append(Paragraph(f"Tại kho: <b>{location_name}</b>, chúng tôi gồm có:", style_meta))
                story_pdf.append(Paragraph(f"Họ tên : <b>{mgr_name}</b> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Chức danh: Quản lý Đơn vị", style_meta))
                story_pdf.append(Paragraph(f"Họ tên : <b>{staff_name}</b> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; Chức danh: Đại diện Kiểm kê / Kho", style_meta))
                story_pdf.append(Spacer(1, 10))

                # Table Header
                table_data_pdf = [[
                    Paragraph("<b>STT</b>", style_tbl_hdr),
                    Paragraph("<b>SKU</b>", style_tbl_hdr),
                    Paragraph("<b>Tên hàng hóa</b>", style_tbl_hdr),
                    Paragraph("<b>ĐVT</b>", style_tbl_hdr),
                    Paragraph("<b>Số lượng sổ sách</b>", style_tbl_hdr),
                    Paragraph("<b>Số lượng thực tế</b>", style_tbl_hdr),
                    Paragraph("<b>Chênh lệch</b>", style_tbl_hdr),
                    Paragraph("<b>Ghi chú/ Tình trạng hàng hóa</b>", style_tbl_hdr)
                ]]

                sum_sys_pdf = df_th_out['Số lượng sổ sách (hệ thống)'].sum()
                sum_act_pdf = df_th_out['Số lượng thực tế'].sum()
                sum_diff_pdf = df_th_out['Chênh lệch'].sum()

                for _, row_p in df_th_out.iterrows():
                    d_val = row_p['Chênh lệch']
                    d_str = f"({abs(d_val)})" if d_val < 0 else ("-" if d_val == 0 else str(d_val))
                    cell_style_diff = style_tbl_cell_red if d_val != 0 else style_tbl_cell

                    # Ghi chú đã lọc sạch thông tin đếm bù nội bộ Lần 2 (ẢNH 2)
                    clean_note_pdf = clean_export_note(row_p['Ghi chú/ Tình trạng hàng hóa'])

                    table_data_pdf.append([
                        Paragraph(str(row_p['STT']), style_tbl_cell),
                        Paragraph(str(row_p['SKU']), style_tbl_cell),
                        Paragraph(str(row_p['Tên hàng hóa']), style_tbl_cell),
                        Paragraph(str(row_p['Đơn vị tính']), style_tbl_cell),
                        Paragraph(f"{row_p['Số lượng sổ sách (hệ thống)']:,}", style_tbl_cell),
                        Paragraph(f"{row_p['Số lượng thực tế']:,}", style_tbl_cell),
                        Paragraph(d_str, cell_style_diff),
                        Paragraph(clean_note_pdf, style_tbl_cell)
                    ])

                # Dòng tổng cộng
                sum_diff_str = f"({abs(sum_diff_pdf)})" if sum_diff_pdf < 0 else str(sum_diff_pdf)
                table_data_pdf.append([
                    Paragraph("<b>Tổng cộng</b>", style_tbl_cell_red),
                    Paragraph("", style_tbl_cell),
                    Paragraph("", style_tbl_cell),
                    Paragraph("", style_tbl_cell),
                    Paragraph(f"<b>{sum_sys_pdf:,}</b>", style_tbl_cell_red),
                    Paragraph(f"<b>{sum_act_pdf:,}</b>", style_tbl_cell_red),
                    Paragraph(f"<b>{sum_diff_str}</b>", style_tbl_cell_red),
                    Paragraph("", style_tbl_cell)
                ])

            col_widths = [30, 70, 230, 40, 80, 80, 70, 200]
            pdf_table = Table(table_data_pdf, colWidths=col_widths, repeatRows=1)
            pdf_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9E1F2')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#AAAAAA')),
                ('SPAN', (0, -1), (3, -1)),
            ]))

            story_pdf.append(pdf_table)
            story_pdf.append(Spacer(1, 15))

            # Footer Chữ ký
            story_pdf.append(Paragraph("<i>Các bên cùng nhau kiểm kê số lượng hàng hóa hư hỏng (nếu có) theo bảng dưới đây:</i>", style_meta))
            story_pdf.append(Spacer(1, 5))
            story_pdf.append(Paragraph(f"<i>Biên bản kết thúc vào lúc __ giờ __ phút, ngày {report_date.strftime('%d')} tháng {report_date.strftime('%m')} năm {report_date.strftime('%Y')}, các thành viên đều thống nhất với nội dung trên.</i>", style_meta))
            story_pdf.append(Paragraph("<i>Biên bản được lập thành ... bản, có giá trị pháp lý như nhau, mỗi bên giữ 01 (một) bản.</i>", style_meta))
            story_pdf.append(Spacer(1, 15))

            sig_table_data = [
                [
                    Paragraph("<b>Đại diện Kho</b>", style_sig_title),
                    Paragraph("<b>Đại diện kiểm kê</b>", style_sig_title),
                    Paragraph("<b>Quản lý Đơn vị</b>", style_sig_title),
                    Paragraph("<b>Kiểm soát nội bộ</b>", style_sig_title)
                ],
                [
                    Paragraph("<i>(Ký, ghi rõ họ tên)</i>", style_sig_sub),
                    Paragraph("<i>(Ký, ghi rõ họ tên)</i>", style_sig_sub),
                    Paragraph("<i>(Ký, ghi rõ họ tên)</i>", style_sig_sub),
                    Paragraph("<i>(Ký, ghi rõ họ tên)</i>", style_sig_sub)
                ]
            ]
            sig_table = Table(sig_table_data, colWidths=[200, 200, 200, 200])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            story_pdf.append(sig_table)

            doc_pdf.build(story_pdf)
            pdf_bytes = pdf_output.getvalue()

            st.download_button(
                label="📄 Tải xuống Biên Bản PDF (Chứa Logo & Lọc sạch Note Lần 2)",
                data=pdf_bytes,
                file_name=f"Bien_Ban_Kiem_Ke_Hang_Ton_Kho_{report_date.strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                type="primary"
            )

    else:
        st.warning("⚠️ Không có dữ liệu để xuất báo cáo. Vui lòng hoàn thành ở các Tab trước.")
